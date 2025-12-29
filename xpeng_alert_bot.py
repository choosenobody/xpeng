#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xpeng_alert_bot.py  (v2.2.3)

本次改动（对应你的两点需求）：
1) 解决 Robotics Rev Share (%) 的 NA：
   - 优先从 KPI_Monitor 中读取（支持多种别名/关键词）
   - 若 KPI_Monitor 没有该行，则支持用环境变量提供数值：
        ROBOTICS_LATEST=6.5   （百分比数值）
        ROBOTICS_TARGET=5     （可选，默认5）
     并在消息中标注“来自环境变量”，避免误解为 Excel 原生数据
   - 可选：AUTO_PATCH_KPI=1 时，若 Robotics 行缺失且提供了 ROBOTICS_LATEST，则自动向 KPI_Monitor 追加该行，避免下次仍 NA

2) 推送内容标题与 KPI 名称全部中文化（代码、现价、整车毛利率等）。

其他：保留 v2.2 的功能，并继续提供 .xlsx 文件有效性检查（ZIP 头 'PK'）以避免 BadZipFile。

环境变量：
- TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID   # 必填(若要推送)
- LIVE_PRICE=1                            # 开启实时股价（默认1）
- YF_SYMBOL=9868.HK                       # 雅虎符号；美股可设 XPEV
- PRICE_FIELD=Close                       # 'Close' 或 'Adj Close'（默认 Close）
- TZ=Asia/Hong_Kong                       # 仅写入日志用
- ROBOTICS_LATEST=6.5                     # Robotics 收入占比（%），用于补齐 Excel 缺失
- ROBOTICS_TARGET=5                       # Robotics 阈值（%），可选
- AUTO_PATCH_KPI=1                        # 可选：自动把 Robotics 行写回 KPI_Monitor
"""

import os, sys, re, csv, datetime
from pathlib import Path
from typing import Optional, Tuple, Dict, Any

import pandas as pd
import numpy as np

# ------------------------- Excel 安全检查 & 读取 -------------------------

def _read_head(path: Path, n: int = 256) -> bytes:
    return path.read_bytes()[:n]

def _diagnose_not_xlsx(head: bytes) -> str:
    text = head.decode("utf-8", "ignore").strip()
    if "git-lfs.github.com/spec/v1" in text:
        return (
            "检测到这是 Git LFS 指针文件，而不是实际的 .xlsx 二进制。\n"
            "修复：GitHub Actions 的 actions/checkout 增加 `lfs: true`，并运行 `git lfs pull`。"
        )
    low = text.lower()
    if low.startswith("<!doctype html") or low.startswith("<html"):
        return (
            "检测到文件内容像 HTML（很可能下载到了 404/鉴权/重定向页面），并非 .xlsx。\n"
            "修复：下载时使用 curl -fL，并检查 URL / 权限 / 重定向。"
        )
    return (
        "文件不是有效的 .xlsx（缺少 ZIP 头 'PK'），可能已损坏或被错误内容覆盖。\n"
        "建议：重新生成/上传该 xlsx，或检查 CI 中的下载与缓存流程。"
    )

def ensure_xlsx_ok(xlsx_path: str) -> None:
    p = Path(xlsx_path)
    if not p.exists():
        raise FileNotFoundError(f"Excel 文件不存在：{xlsx_path}")
    head2 = _read_head(p, 2)
    if head2 != b"PK":
        head = _read_head(p, 256)
        raise ValueError(
            f"Excel 文件不是有效 .xlsx：{xlsx_path}\n{_diagnose_not_xlsx(head)}\n"
            f"文件大小：{p.stat().st_size} bytes"
        )

def read_sheet_safe(xlsx_path: str, sheet: str) -> pd.DataFrame:
    ensure_xlsx_ok(xlsx_path)
    return pd.read_excel(xlsx_path, sheet_name=sheet, engine="openpyxl")

# ------------------------- Yahoo 价格 -------------------------

def fetch_live_price(symbol: str, price_field: str = "Close") -> Optional[float]:
    try:
        import yfinance as yf
        t = yf.Ticker(symbol)
        df = t.history(period="1d")
        if df.shape[0] == 0:
            return None
        field = price_field if price_field in df.columns else "Close"
        px = float(df[field].iloc[-1])
        if np.isnan(px):
            return None
        return px
    except Exception:
        return None

# ------------------------- Excel 读写 -------------------------

from openpyxl import load_workbook

def update_assumptions_price(xlsx_path: str, new_price: float) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb["Assumptions"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    item_idx = headers.index("Item") + 1
    val_idx  = headers.index("Value") + 1
    found = False
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, item_idx).value).strip() == "Current Price":
            ws.cell(r, val_idx, float(new_price))
            found = True
            break
    if not found:
        ws.append(["Current Price", float(new_price), "HKD", "auto-updated"])

    if "Status_Log" not in wb.sheetnames:
        wb.create_sheet("Status_Log")
        wsl = wb["Status_Log"]
        wsl.append([
            "timestamp_utc","price_hkd","base_iv_hkd","discount_pct",
            "ok_vehicle_gm","ok_fcf","ok_techsvc","ok_robotics",
            "kpi_pass","signal","rating_upgrade"
        ])
    wb.save(xlsx_path)

def maybe_patch_kpi_robotics_row(xlsx_path: str, latest: float, target: float) -> bool:
    """
    AUTO_PATCH_KPI=1 时：若 KPI_Monitor 缺失 Robotics 行，则追加一行。
    仅在 KPI_Monitor 结构为 [Metric, Latest, Target/Threshold] 时启用。
    返回是否成功写入。
    """
    if os.environ.get("AUTO_PATCH_KPI", "0") != "1":
        return False
    try:
        wb = load_workbook(xlsx_path)
        if "KPI_Monitor" not in wb.sheetnames:
            return False
        ws = wb["KPI_Monitor"]
        headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value is not None else "" for c in range(1, ws.max_column+1)]
        if "Metric" not in headers or "Latest" not in headers or "Target/Threshold" not in headers:
            return False
        mcol = headers.index("Metric") + 1
        lcol = headers.index("Latest") + 1
        tcol = headers.index("Target/Threshold") + 1

        # 是否已存在（含关键词匹配）
        for r in range(2, ws.max_row+1):
            v = ws.cell(r, mcol).value
            if v is None:
                continue
            s = str(v)
            if ("robot" in s.lower()) or ("机器人" in s) or ("Robotics" in s):
                # 已有行就不重复写
                return False

        ws.append(["Robotics Rev Share (%)", float(latest), float(target)])
        wb.save(xlsx_path)
        return True
    except Exception:
        return False

# ------------------------- DCF 兜底（若 Summary 缺失） -------------------------

def compute_wacc(rf, erp, beta, tax, debt_ratio, pre_tax_cost_debt):
    ke = rf + beta * erp
    kd_after = pre_tax_cost_debt * (1 - tax)
    return ke * (1 - debt_ratio) + kd_after * debt_ratio

def project_revenue_series(start_rev, cagr, n_years=10):
    return [start_rev * ((1 + cagr) ** i) for i in range(1, n_years+1)]

def dcf_base_iv(xlsx_path: str) -> Optional[float]:
    try:
        A = read_sheet_safe(xlsx_path, "Assumptions")
        R = read_sheet_safe(xlsx_path, "Start_Rev_2025")
        S = read_sheet_safe(xlsx_path, "Scenarios")
        amap = dict(zip(A["Item"], A["Value"]))
        rf=float(amap.get("Risk-Free Rate (Rf)",0.0181))
        erp=float(amap.get("Equity Risk Premium (ERP)",0.059))
        beta=float(amap.get("Beta",1.04))
        tax=float(amap.get("Tax Rate",0.25))
        d_ratio=float(amap.get("Target Debt Ratio (D/(D+E))",0.10))
        kd_pre=float(amap.get("Pre-tax Cost of Debt",0.045))
        g=float(amap.get("Terminal Growth (g)",0.02))
        s2c=float(amap.get("Sales-to-Capital",2.5))
        shares=float(amap.get("Share Count (bn)",1.909771413))
        net_cash=float(amap.get("Net Cash (bn)",39.9))
        start_rev=float(R["Value"].iloc[0])
        wacc = compute_wacc(rf, erp, beta, tax, d_ratio, kd_pre)

        base_df = S[S["Scenario"]=="Base"].copy()
        rev_cagr = float(base_df["Rev_CAGR"].iloc[0])
        ebit_path = base_df["EBIT_margin"].values.astype(float)

        rev = np.array(project_revenue_series(start_rev, rev_cagr, n_years=len(ebit_path)))
        ebit = rev * ebit_path
        nopat = ebit * (1 - tax)
        reinv = (rev * rev_cagr) / max(1e-6, s2c)
        fcff = nopat - reinv

        years = np.arange(1, len(fcff)+1)
        disc = (1 + wacc) ** years
        pv_fcff = float(np.sum(fcff / disc))
        tv = float((fcff[-1] * (1 + g)) / (wacc - g))
        pv_tv = float(tv / ((1+wacc)**len(fcff)))
        ev = pv_fcff + pv_tv
        equity = ev + net_cash
        per_share = (equity * 1e9) / (shares * 1e9)
        return float(per_share)
    except Exception:
        return None

# ------------------------- KPI 解析（增强：输出 Latest/Target/原因；支持 env 补齐 Robotics） -------------------------

def _to_float(x) -> Optional[float]:
    try:
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return None
        s = str(x).strip()
        if s == "":
            return None
        return float(s)
    except Exception:
        return None

def _parse_target(x, default: Optional[float] = None) -> Optional[float]:
    if x is None:
        return default
    s = str(x)
    m = re.search(r"(-?\d+(\.\d+)?)", s)
    if not m:
        return default
    try:
        return float(m.group(1))
    except Exception:
        return default

def _get_metric_row(K: pd.DataFrame, names: list, contains_keywords: list = None):
    mcol = K["Metric"].astype(str).str.strip()
    for name in names:
        row = K[mcol == str(name).strip()]
        if not row.empty:
            return row.iloc[0]
    if contains_keywords:
        for kw in contains_keywords:
            row = K[mcol.str.contains(str(kw), case=False, na=False)]
            if not row.empty:
                return row.iloc[0]
    return None

def _eval_kpi_ge(row, default_target: float) -> Tuple[Optional[bool], Optional[float], Optional[float], str]:
    if row is None:
        return None, None, None, "KPI_Monitor 未提供该指标行"
    latest = _to_float(row.get("Latest"))
    target = _parse_target(row.get("Target/Threshold"), default_target)
    if latest is None:
        return None, None, target, "Latest 为空/不可解析"
    if target is None:
        return None, latest, None, "Target/Threshold 为空/不可解析"
    ok = bool(latest >= target)
    return ok, float(latest), float(target), ""

def kpi_details(K: pd.DataFrame) -> Dict[str, Any]:
    gm_row = _get_metric_row(K,
                            ["Vehicle GM (%)", "Vehicle GM", "Vehicle GM%"],
                            ["Vehicle GM", "GM", "整车毛利", "毛利率"])
    fcf_row = _get_metric_row(K,
                             ["FCF (TTM, bn HKD)", "FCF (TTM)", "FCF"],
                             ["FCF", "自由现金流"])
    ts_row  = _get_metric_row(K,
                             ["Tech/Service Rev Share (%)", "Tech/Service Share (%)", "Tech/Service"],
                             ["Tech", "Service", "服务", "科技"])
    rb_row  = _get_metric_row(K,
                             ["Robotics Rev Share (%)", "Robotics Share (%)", "Robotics"],
                             ["robot", "机器人", "robotics"])

    ok_gm,  gm_latest, gm_target, gm_reason = _eval_kpi_ge(gm_row, 15)
    ok_fcf, fcf_latest, fcf_target, fcf_reason = _eval_kpi_ge(fcf_row, 0)
    ok_ts,  ts_latest, ts_target, ts_reason = _eval_kpi_ge(ts_row, 10)

    # Robotics：先读表；缺失则看环境变量
    rb_source = "excel"
    if rb_row is None:
        env_latest = _to_float(os.environ.get("ROBOTICS_LATEST"))
        if env_latest is not None:
            env_target = _to_float(os.environ.get("ROBOTICS_TARGET"))
            if env_target is None:
                env_target = 5.0
            ok_rb = bool(env_latest >= env_target)
            rb_latest, rb_target, rb_reason = float(env_latest), float(env_target), "来自环境变量 ROBOTICS_LATEST"
            rb_source = "env"
            # 可选自动补写 KPI_Monitor
            maybe_patch_kpi_robotics_row(os.path.abspath(xlsx_path_global), rb_latest, rb_target)  # noqa
        else:
            ok_rb, rb_latest, rb_target, rb_reason = None, None, 5.0, "KPI_Monitor 未提供 Robotics 指标行，且未设置 ROBOTICS_LATEST"
    else:
        ok_rb, rb_latest, rb_target, rb_reason = _eval_kpi_ge(rb_row, 5)

    # “机器人/技术服务”达标
    if ok_ts is None and ok_rb is None:
        ok_rt = None
    else:
        ok_rt = bool(ok_ts is True or ok_rb is True)

    kpi_pass = 0
    if ok_gm is True: kpi_pass += 1
    if ok_fcf is True: kpi_pass += 1
    if ok_rt is True: kpi_pass += 1

    return dict(
        ok_gm=ok_gm, gm_latest=gm_latest, gm_target=gm_target, gm_reason=gm_reason,
        ok_fcf=ok_fcf, fcf_latest=fcf_latest, fcf_target=fcf_target, fcf_reason=fcf_reason,
        ok_ts=ok_ts, ts_latest=ts_latest, ts_target=ts_target, ts_reason=ts_reason,
        ok_rb=ok_rb, rb_latest=rb_latest, rb_target=rb_target, rb_reason=rb_reason, rb_source=rb_source,
        ok_rt=ok_rt,
        kpi_pass=kpi_pass
    )

def _fmt(x: Optional[float], nd: int = 2) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return "NA"
    try:
        return f"{float(x):.{nd}f}"
    except Exception:
        return "NA"

def _pf(ok: Optional[bool]) -> str:
    if ok is True: return "PASS"
    if ok is False: return "FAIL"
    return "NA"

# ------------------------- 状态记录 -------------------------

def append_logs(xlsx_path, price, base_iv, ok_gm, ok_fcf, ok_ts, ok_rb, ok_rt, kpi_pass, signal, rating_up):
    ts_utc = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    discount = (price/base_iv - 1.0)*100 if (base_iv and base_iv==base_iv and base_iv>0) else np.nan

    row = {
        "timestamp_utc": ts_utc,
        "price_hkd": round(price, 4) if price==price else "",
        "base_iv_hkd": round(base_iv, 4) if base_iv==base_iv else "",
        "discount_pct": round(discount, 3) if discount==discount else "",
        "ok_vehicle_gm": int(ok_gm is True),
        "ok_fcf": int(ok_fcf is True),
        "ok_techsvc": int(ok_ts is True),
        "ok_robotics": int(ok_rb is True),
        "kpi_pass": int(kpi_pass),
        "signal": signal,
        "rating_upgrade": int(rating_up is True)
    }
    csv_path = "status_log.csv"
    write_header = not os.path.exists(csv_path)
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(row.keys()))
        if write_header:
            w.writeheader()
        w.writerow(row)

    try:
        wb = load_workbook(xlsx_path)
        ws = wb["Status_Log"] if "Status_Log" in wb.sheetnames else wb.create_sheet("Status_Log")
        if ws.max_row == 1 and ws.cell(1,1).value != "timestamp_utc":
            ws.append(["timestamp_utc","price_hkd","base_iv_hkd","discount_pct",
                       "ok_vehicle_gm","ok_fcf","ok_techsvc","ok_robotics",
                       "kpi_pass","signal","rating_upgrade"])
        ws.append([ts_utc, price, base_iv, discount,
                   int(ok_gm is True), int(ok_fcf is True),
                   int(ok_ts is True), int(ok_rb is True),
                   int(kpi_pass), signal, int(rating_up is True)])
        wb.save(xlsx_path)
    except Exception:
        pass

# ------------------------- Telegram -------------------------

def send_telegram(text: str):
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        print("TELEGRAM_BOT_TOKEN/TELEGRAM_CHAT_ID 未配置；仅打印：\n"+text)
        return
    import urllib.request, urllib.parse
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    data = urllib.parse.urlencode({
        "chat_id": chat_id,
        "text": text,
        "parse_mode": "Markdown"
    }).encode("utf-8")
    with urllib.request.urlopen(url, data=data, timeout=20) as r:
        r.read()

# ------------------------- 主流程 -------------------------

xlsx_path_global = ""  # 给 kpi_details 内部 auto patch 用

def main(xlsx_path: str):
    global xlsx_path_global
    xlsx_path_global = xlsx_path

    try:
        ensure_xlsx_ok(xlsx_path)
    except Exception as e:
        send_telegram(f"📉 小鹏监控：Excel 文件不可用\n\n{e}")
        return 0

    # 1) 读取 Assumptions
    try:
        A = read_sheet_safe(xlsx_path, "Assumptions")
        amap = dict(zip(A["Item"], A["Value"]))
    except Exception as e:
        send_telegram(f"📉 小鹏监控：读取 Assumptions 失败\n\n{e}")
        return 0

    # 2) 实时股价（可选）
    live = os.environ.get("LIVE_PRICE","1") == "1"
    symbol = os.environ.get("YF_SYMBOL","9868.HK")
    price_field = os.environ.get("PRICE_FIELD","Close")
    price_live = fetch_live_price(symbol, price_field) if live else None
    price = float(price_live) if (price_live is not None) else float(amap.get("Current Price", 0))

    if price_live is not None:
        try:
            update_assumptions_price(xlsx_path, price)
        except Exception as e:
            send_telegram(f"⚠️ 小鹏监控：实时价格写回失败（不影响本次信号计算）\n\n{e}")

    # 3) Base IV：优先 Summary；否则 DCF
    base_iv = None
    try:
        S = read_sheet_safe(xlsx_path, "Summary")
        base_row = S[S["Scenario"]=="Base"]
        base_iv = float(base_row["IV_HKD_per_share"].values[0]) if not base_row.empty else None
    except Exception:
        base_iv = None
    if (base_iv is None) or (base_iv != base_iv):
        base_iv = dcf_base_iv(xlsx_path)

    # 4) KPI
    try:
        K = read_sheet_safe(xlsx_path, "KPI_Monitor")
    except Exception as e:
        send_telegram(f"📉 小鹏监控：读取 KPI_Monitor 失败\n\n{e}")
        return 0

    kd = kpi_details(K)
    ok_gm, ok_fcf, ok_ts, ok_rb, ok_rt, kpi_pass = (
        kd["ok_gm"], kd["ok_fcf"], kd["ok_ts"], kd["ok_rb"], kd["ok_rt"], kd["kpi_pass"]
    )

    # 5) 信号 & 评级建议
    signal = "观察"
    if base_iv and base_iv==base_iv and base_iv > 0:
        if price <= 0.80 * base_iv:
            signal = "加仓"
        elif price <= 0.90 * base_iv:
            signal = "建仓"
    rating_up = (kpi_pass >= 2) and (ok_rt is True)

    # 6) 记录
    append_logs(xlsx_path, price, base_iv, ok_gm, ok_fcf, ok_ts, ok_rb, ok_rt, kpi_pass, signal, rating_up)

    # 7) 中文化推送
    ts_utc = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    if base_iv and base_iv==base_iv and base_iv > 0:
        premium_pct = (price / base_iv - 1.0) * 100
        iv_line = f"基准内在价值: HK${base_iv:.2f} | 溢价: {premium_pct:+.1f}%"
    else:
        iv_line = "基准内在价值: N/A"

    lines = []
    lines.append("*小鹏估值监控*")
    lines.append(f"时间: {ts_utc}")
    lines.append(f"代码: `{symbol}` | 现价: HK${price:.2f}")
    lines.append(iv_line)
    lines.append(f"信号: *{signal}* | KPI通过数: {kpi_pass}/3 | 评级建议: {'*上调*' if rating_up else '暂不升级'}")
    lines.append("")
    lines.append("*KPI 明细（最新值 vs 阈值 → 结论）*")

    # 整车毛利率
    if kd["ok_gm"] is None:
        lines.append(f"- 整车毛利率(%): NA（{kd['gm_reason']}）")
    else:
        lines.append(f"- 整车毛利率(%): {_fmt(kd['gm_latest'])} vs ≥{_fmt(kd['gm_target'])} → {_pf(kd['ok_gm'])}")

    # 自由现金流
    if kd["ok_fcf"] is None:
        lines.append(f"- 自由现金流TTM(十亿港币): NA（{kd['fcf_reason']}）")
    else:
        lines.append(f"- 自由现金流TTM(十亿港币): {_fmt(kd['fcf_latest'])} vs ≥{_fmt(kd['fcf_target'])} → {_pf(kd['ok_fcf'])}")

    # 科技/服务收入占比
    if kd["ok_ts"] is None:
        lines.append(f"- 科技/服务收入占比(%): NA（{kd['ts_reason']}）")
    else:
        lines.append(f"- 科技/服务收入占比(%): {_fmt(kd['ts_latest'])} vs ≥{_fmt(kd['ts_target'])} → {_pf(kd['ok_ts'])}")

    # 机器人收入占比
    if kd["ok_rb"] is None:
        lines.append(f"- 机器人收入占比(%): NA（{kd['rb_reason']}；你可以设置 ROBOTICS_LATEST=xx 以补齐）")
    else:
        src = "（来自环境变量）" if kd.get("rb_source") == "env" else ""
        lines.append(f"- 机器人收入占比(%): {_fmt(kd['rb_latest'])} vs ≥{_fmt(kd['rb_target'])} → {_pf(kd['ok_rb'])} {src}")

    # 综合
    if ok_rt is None:
        rt_line = "NA（科技/服务 与 机器人 均缺失）"
    else:
        rt_line = _pf(ok_rt)
    lines.append(f"- 机器人/科技服务综合（任一PASS即PASS）：{rt_line}")

    send_telegram("\n".join(lines))
    return 0

if __name__=="__main__":
    if len(sys.argv)<2:
        print("Usage: python xpeng_alert_bot.py /path/to/XPeng_Valuation_Monitor_v2.xlsx")
        sys.exit(1)
    sys.exit(main(sys.argv[1]))
